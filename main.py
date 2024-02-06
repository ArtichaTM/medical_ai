import pandas as pd
from openpyxl import load_workbook, Workbook
from matplotlib import pyplot as plt


def load_info() -> pd.DataFrame:
    wb: Workbook = load_workbook(filename='ПАТТЕРН ОБЕЗЛИЧ..xlsx')
    columns = [
        'sex', 'age',
        'IN T', 'IN P', 'IN P1/P', 'IN P2/P', 'IN P3/P', 'IN P1', 'IN P2', 'IN P3',
        'OUT T', 'OUT P', 'OUT P1/P', 'OUT P2/P', 'OUT P3/P', 'OUT P1', 'OUT P2', 'OUT P3',
        'SUM T', 'SUM P', 'SUM P1/P', 'SUM P2/P', 'SUM P3/P', 'SUM P1', 'SUM P2', 'SUM P3',
        'diagnose'
    ]
    iterator = wb.worksheets[0]
    iterator = (row for row in iterator.iter_rows(min_row=3, min_col=0, max_col=len(columns)))
    iterator = ([i.value for i in row] for row in iterator)
    df = pd.DataFrame(iterator, columns=columns)
    return df


def calculate_line(df: pd.DataFrame):
    all_diagnoses = df['diagnose'].unique()
    for name in df.columns:
        if any((
            name in {'sex', 'age', 'diagnose'},
            '/' in name
        )):
            continue
        df.sort_values(name, inplace=True, ignore_index=True)
        f, ax = plt.subplots(figsize=(19.20, 10.80))
        diagnoses = {i: ([], []) for i in all_diagnoses}
        for row_id, series in df.iterrows():
            if series[name] is None:
                continue
            lists = diagnoses[series['diagnose']]
            lists[0].append(row_id)
            lists[1].append(series[name])
        for diagnose, (x, y) in diagnoses.items():
            ax.scatter(x, y, label=diagnose, s=11)

        ax: plt.Axes
        ax.set_title(name)
        name = name.replace(' ', '_').replace('/', '_')
        plt.legend(all_diagnoses)
        plt.savefig(f"graphs\\line\\{name}.png")
        plt.close('all')


def calculate_diagnose_bars(df: pd.DataFrame) -> pd.Series:
    plt.close()
    ax = plt.subplot()
    values = df['diagnose'].value_counts()
    ax.bar(x=values.index, height=values.values)
    ax.set_ylabel('Количество')
    ax.set_xlabel('Диагноз')
    plt.setp(ax.get_xticklabels(), rotation=10, ha='right')
    plt.tight_layout()
    plt.show()
    return values


def calculate_density(df: pd.DataFrame):
    for name in df.columns:
        ax = df[name].plot(kind='kde')
        ax: plt.Axes
        ax.set_title(name)
        name = name.replace(' ', '_').replace('/', '_')
        plt.savefig(f"graphs\\density\\{name}.png")
        plt.close('all')


def main():
    df = load_info()
    # calculate_diagnose_bars(df)
    calculate_line(df)
    # print(df.info())


if __name__ == '__main__':
    main()
