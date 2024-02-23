import pandas as pd
import keras
import tensorflow as tf
from keras.utils import FeatureSpace
from openpyxl import load_workbook, Workbook


import os
os.environ["KERAS_BACKEND"] = "tensorflow"
os.environ['TF_CPP_MIN_LOG_LEVEL'] = '3'


__all__ = (
    'load_info',
    'drop_other_diagnoses',
    'diagnoses_to_digits',
    'dataframe_to_dataset',
    'prepare_datasets',
    'build_feature_space',
    'build_model',
)


def build_feature_space():
    return FeatureSpace(
        features={
            'sex'   : FeatureSpace.integer_categorical(),
            'age'   : FeatureSpace.float_rescaled(),
            'IN T'  : FeatureSpace.float_normalized(),
            'IN P'  : FeatureSpace.float_normalized(),
            'IN P1' : FeatureSpace.float_normalized(),
            'IN P2' : FeatureSpace.float_normalized(),
            'IN P3' : FeatureSpace.float_normalized(),
            'OUT T' : FeatureSpace.float_normalized(),
            'OUT P'  : FeatureSpace.float_normalized(),
            'OUT P1': FeatureSpace.float_normalized(),
            'OUT P2': FeatureSpace.float_normalized(),
            'OUT P3': FeatureSpace.float_normalized(),
        },
        output_mode='concat'
    )


def drop_other_diagnoses(
    data: pd.DataFrame,
    exclude_columns: list[str] = None
) -> pd.DataFrame:
    indexes = []
    for index, row in enumerate(data['diagnose']):
        if row not in {'норма', 'астма ремиссия'}:
            indexes.append(index)
    data.drop(indexes, inplace=True)

    if exclude_columns is not None:
        data.drop(columns=exclude_columns, inplace=True)

    feature_space = build_feature_space()
    features = set(feature_space.features)
    features.add('diagnose')
    features = set(data.columns).difference(features)
    if features:
        data.drop(columns=list(features), inplace=True)

    return data


def diagnoses_to_digits(df: pd.DataFrame) -> None:
    diagnoses = df['diagnose'].unique()
    dictionary = {diagnose: index for index, diagnose in enumerate(diagnoses)}
    df.replace({'diagnose': dictionary}, inplace=True)


def dataframe_to_dataset(dataframe: pd.DataFrame) -> tf.data.Dataset:
    dataframe = dataframe.copy()
    labels = dataframe.pop("diagnose")
    ds = tf.data.Dataset.from_tensor_slices((dict(dataframe), labels))
    ds = ds.shuffle(buffer_size=len(dataframe))
    return ds


def prepare_datasets(
    data: pd.DataFrame,
    sample_fraction: float
) -> tuple[tf.data.Dataset, tf.data.Dataset]:
    """
    :return: Train and validation dataset
    """
    df_validation = data.sample(frac=sample_fraction, random_state=1658179)
    df_train = data.drop(df_validation.index)
    ds_validation = dataframe_to_dataset(df_validation)
    ds_train = dataframe_to_dataset(df_train)
    ds_validation = ds_validation.batch(32)
    ds_train = ds_train.batch(32)
    return ds_train, ds_validation



def load_info() -> pd.DataFrame:
    wb: Workbook = load_workbook(filename='ПАТТЕРН ОБЕЗЛИЧ..xlsx')
    columns = [
        'sex', 'age',
        'IN T', 'IN P', 'IN P1/P', 'IN P2/P', 'IN P3/P', 'IN P1', 'IN P2', 'IN P3',
        'OUT T', 'OUT P', 'OUT P1/P', 'OUT P2/P', 'OUT P3/P', 'OUT P1', 'OUT P2', 'OUT P3',
        'SUM T', 'SUM P', 'SUM P1/P', 'SUM P2/P', 'SUM P3/P', 'SUM P1', 'SUM P2', 'SUM P3',
        'diagnose'
    ]
    iterator = wb.worksheets[0]
    iterator = (row for row in iterator.iter_rows(min_row=3, min_col=0, max_col=len(columns)))
    iterator = ([i.value for i in row] for row in iterator)
    df = pd.DataFrame(iterator, columns=columns)
    return df


def build_model(inputs: int) -> keras.Model:
    model = keras.Sequential(
        layers=[
            keras.layers.Input((inputs,)),
            # keras.layers.Dense(inputs * 4),
            # keras.layers.Dropout(0.1),
            keras.layers.Dense(inputs * 6),
            # keras.layers.Dropout(0.2),
            keras.layers.Dropout(0.5),
            # keras.layers.Dense(inputs * 4),
            # keras.layers.Dropout(0.1),
            keras.layers.Dense(1, activation='sigmoid')
        ]
    )
    model.build()
    model.compile(optimizer="adam")
    return model
